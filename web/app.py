"""
Flask Web 应用 — 数据挖掘与机器学习平台
支持: 数据集浏览/上传、算法参数配置、训练评估、多维可视化、报告导出
"""
import os
import sys
import json
import time
import io
import uuid
import base64
import traceback
import csv
import hashlib

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import numpy as np
from flask import Flask, render_template, request, jsonify, send_file

from config import WEB_CONFIG, DATASETS, ALGORITHM_CONFIG, PREPROCESSING, OUTPUT_DIR
from src.data_loader import load_dataset
from src.preprocessing import StandardScaler, MinMaxScaler, train_test_split
from src.evaluation import cross_validate, compute_all_metrics, k_fold_split

# 算法模块
from src.algorithms.knn import KNN
from src.algorithms.naive_bayes import GaussianNB
from src.algorithms.decision_tree import DecisionTreeC45
from src.algorithms.svm import SVM
from src.algorithms.kmeans import KMeans
from src.algorithms.mlp import MLP

# 可视化模块
from src.visualization import (
    plot_histograms, plot_boxplots, plot_violin, plot_swarm,
    plot_correlation_heatmap, plot_scatter_matrix,
    plot_pca_2d, plot_pca_3d, plot_pca_explained_variance,
    plot_parallel_coordinates, plot_radar, plot_dendrogram,
    plot_ecdf, plot_class_pie, plot_feature_importance,
    plot_pair_density, plot_learning_curve,
    plot_confusion_matrix as viz_confusion_matrix,
    plot_algorithm_comparison, plot_time_comparison,
    generate_html_report, OUTPUT_DIR as VIZ_OUTPUT_DIR,
)

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

# ── 中文字体配置 ──────────────────────────────────────
plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False

app = Flask(__name__)
app.config["SECRET_KEY"] = "ml-platform-secret-key-2026"
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024  # 100MB 上传限制

# ── 静态文件路由：output 目录 ─────────────────────────
from flask import send_from_directory as _send_from_directory
_APP_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

@app.route("/output/<path:filename>")
def serve_output(filename):
    """直接提供 output 目录下的静态文件（图表等）"""
    return _send_from_directory(os.path.join(_APP_ROOT, "output"), filename)

# 上传目录
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                              WEB_CONFIG.get("upload_folder", "uploads"))
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(VIZ_OUTPUT_DIR, exist_ok=True)

# 结果缓存: {run_id: {dataset, algo, metrics, chart_b64, ...}}
results_cache = {}
# 上传数据缓存: {upload_id: {X, y, feature_names, class_names, filename, ...}}
upload_cache = {}
_run_counter = [0]


# ============================================================
# 页面路由
# ============================================================

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/health")
def health():
    return {"status": "ok", "timestamp": time.time()}


# ============================================================
# 数据集 API
# ============================================================

@app.route("/api/datasets")
def api_datasets():
    """返回所有可用数据集元信息"""
    result = {}
    for key, cfg in DATASETS.items():
        result[key] = {
            "name": cfg["name"],
            "n_classes": cfg.get("n_classes", "?"),
            "n_features": cfg.get("n_features", "?"),
            "task": cfg.get("task", "classification"),
        }
    return jsonify(result)


@app.route("/api/datasets/<name>/load")
def api_dataset_load(name):
    """加载并返回数据集摘要"""
    try:
        X, y, fnames, cnames = load_dataset(name)
        n_samples, n_features = X.shape
        n_classes = len(np.unique(y))

        # 每列基本统计
        stats = []
        for i in range(min(n_features, 10)):
            col = X[:, i]
            stats.append({
                "index": i,
                "name": fnames[i] if fnames else f"特征_{i}",
                "mean": round(float(np.mean(col)), 4),
                "std": round(float(np.std(col)), 4),
                "min": round(float(np.min(col)), 4),
                "max": round(float(np.max(col)), 4),
            })

        # 类别分布
        class_counts = {}
        for c in np.unique(y):
            class_counts[str(c)] = int(np.sum(y == c))

        return jsonify({
            "dataset": name,
            "name": DATASETS[name]["name"],
            "n_samples": n_samples,
            "n_features": n_features,
            "n_classes": n_classes,
            "class_names": [str(c) for c in cnames] if cnames else [str(c) for c in np.unique(y)],
            "class_counts": class_counts,
            "feature_stats": stats,
            "missing": int(np.sum(np.isnan(X))),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 400


# ── 基准测试 API ──────────────────────────────────────

@app.route("/api/benchmark")
def api_benchmark():
    """返回完整 Benchmark 数据（5数据集 × 6算法）"""
    json_path = os.path.join(_APP_ROOT, "output", "benchmark_results.json")
    if not os.path.exists(json_path):
        return jsonify({
            "error": "未找到基准测试文件，请先运行: python src/benchmark.py",
            "datasets": {},
        }), 404

    with open(json_path, "r", encoding="utf-8") as f:
        bdata = json.load(f)

    # 构建前端友好的结构化数据
    datasets_info = {}
    for key, cfg in DATASETS.items():
        datasets_info[key] = {
            "name": cfg["name"],
            "n_samples": cfg.get("n_samples", "?"),
            "n_features": cfg.get("n_features", "?"),
            "n_classes": cfg.get("n_classes", "?"),
        }

    algos_order = ["KNN", "Naive Bayes", "C4.5", "SVM", "K-Means", "MLP"]

    # 计算各数据集最优算法
    best_per_dataset = {}
    for ds_key, ds_data in bdata.items():
        best_acc = 0
        best_algo = ""
        for algo_name in algos_order:
            if algo_name in ds_data:
                acc = ds_data[algo_name].get("accuracy", 0)
                if algo_name != "K-Means" and acc > best_acc:
                    best_acc = acc
                    best_algo = algo_name
        if best_algo:
            best_per_dataset[ds_key] = {
                "algorithm": best_algo,
                "accuracy": round(best_acc * 100, 2),
                "dataset_name": datasets_info.get(ds_key, {}).get("name", ds_key),
            }

    return jsonify({
        "datasets_meta": datasets_info,
        "results": bdata,
        "algorithms": algos_order,
        "best_per_dataset": best_per_dataset,
        "has_charts": os.path.exists(os.path.join(_APP_ROOT, "output", "cross_dataset_accuracy.png")),
    })


# ============================================================
# 算法配置 API
# ============================================================

@app.route("/api/algorithms")
def api_algorithms():
    """返回所有算法及参数配置"""
    return jsonify({
        "knn": {
            "name": "KNN",
            "params": {
                "k": {"type": "int", "default": 5, "min": 1, "max": 30, "label": "近邻数 K"},
                "weights": {"type": "select", "default": "uniform", "options": ["uniform", "distance", "gaussian"], "label": "权重策略"},
                "metric": {"type": "select", "default": "euclidean", "options": ["euclidean", "manhattan"], "label": "距离度量"},
                "sigma": {"type": "float", "default": None, "label": "高斯核带宽(自动)"},
            }
        },
        "naive_bayes": {
            "name": "朴素贝叶斯",
            "params": {
                "variant": {"type": "select", "default": "gaussian", "options": ["gaussian"], "label": "变体"},
            }
        },
        "decision_tree": {
            "name": "决策树 C4.5",
            "params": {
                "max_depth": {"type": "int", "default": 10, "min": 1, "max": 50, "label": "最大深度"},
                "min_samples_split": {"type": "int", "default": 2, "min": 2, "max": 20, "label": "最小分裂样本数"},
                "pruning": {"type": "select", "default": "post", "options": ["none", "pre", "post"], "label": "剪枝策略"},
            }
        },
        "svm": {
            "name": "SVM (SMO)",
            "params": {
                "kernel": {"type": "select", "default": "rbf", "options": ["linear", "rbf", "poly"], "label": "核函数"},
                "C": {"type": "float", "default": 1.0, "min": 0.01, "max": 100, "label": "正则化 C"},
                "max_iter": {"type": "int", "default": 5000, "min": 100, "max": 20000, "label": "最大迭代"},
            }
        },
        "kmeans": {
            "name": "K-Means",
            "params": {
                "n_clusters": {"type": "int", "default": 3, "min": 2, "max": 20, "label": "聚类数"},
                "init": {"type": "select", "default": "kmeans++", "options": ["random", "kmeans++"], "label": "初始化方式"},
                "max_iter": {"type": "int", "default": 300, "min": 10, "max": 1000, "label": "最大迭代"},
            }
        },
        "mlp": {
            "name": "MLP 神经网络",
            "params": {
                "hidden_layers": {"type": "string", "default": "64,32", "label": "隐层结构(逗号分隔)"},
                "activation": {"type": "select", "default": "relu", "options": ["relu", "sigmoid", "tanh"], "label": "激活函数"},
                "learning_rate": {"type": "float", "default": 0.001, "min": 0.0001, "max": 0.1, "label": "学习率"},
                "epochs": {"type": "int", "default": 200, "min": 10, "max": 1000, "label": "训练轮数"},
                "batch_size": {"type": "int", "default": 32, "min": 8, "max": 256, "label": "批量大小"},
                "optimizer": {"type": "select", "default": "adam", "options": ["sgd", "momentum", "adam"], "label": "优化器"},
            }
        },
    })


# ============================================================
# 训练 & 评估 API
# ============================================================

def _get_algo_instance(algo_key, params):
    """根据参数创建算法实例"""
    p = params or {}

    if algo_key == "knn":
        return KNN(
            k=int(p.get("k", 5)),
            weights=p.get("weights", "uniform"),
            metric=p.get("metric", "euclidean"),
            sigma=p.get("sigma", None) if p.get("sigma") else None,
        )
    elif algo_key == "naive_bayes":
        return GaussianNB()
    elif algo_key == "decision_tree":
        return DecisionTreeC45(
            max_depth=int(p.get("max_depth", 10)),
            min_samples_split=int(p.get("min_samples_split", 2)),
            pruning=p.get("pruning", "post"),
        )
    elif algo_key == "svm":
        return SVM(
            kernel=p.get("kernel", "rbf"),
            C=float(p.get("C", 1.0)),
            max_iter=int(p.get("max_iter", 5000)),
        )
    elif algo_key == "kmeans":
        return KMeans(
            n_clusters=int(p.get("n_clusters", 3)),
            init=p.get("init", "kmeans++"),
            max_iter=int(p.get("max_iter", 300)),
        )
    elif algo_key == "mlp":
        h_str = p.get("hidden_layers", "64,32")
        try:
            hidden = [int(x.strip()) for x in h_str.split(",") if x.strip()]
        except ValueError:
            hidden = [64, 32]
        return MLP(
            hidden_layers=hidden,
            activation=p.get("activation", "relu"),
            learning_rate=float(p.get("learning_rate", 0.001)),
            epochs=int(p.get("epochs", 200)),
            batch_size=int(p.get("batch_size", 32)),
            optimizer=p.get("optimizer", "adam"),
        )
    return None


@app.route("/api/train", methods=["POST"])
def api_train():
    """训练+评估单个算法"""
    data = request.get_json() or {}
    ds_key = data.get("dataset", "iris")
    algo_key = data.get("algorithm", "knn")
    params = data.get("params", {})
    test_size = float(data.get("test_size", 0.2))
    normalize = data.get("normalize", "standard")
    use_cv = data.get("cross_validation", False)

    try:
        # 加载数据
        X, y, fnames, cnames = load_dataset(ds_key)

        # 标准化
        if normalize == "standard":
            scaler = StandardScaler()
            X = scaler.fit_transform(X)
        elif normalize == "minmax":
            scaler = MinMaxScaler()
            X = scaler.fit_transform(X)

        # 创建算法实例
        model = _get_algo_instance(algo_key, params)
        if model is None:
            return jsonify({"error": f"未知算法: {algo_key}"}), 400

        algo_names = {
            "knn": "KNN", "naive_bayes": "朴素贝叶斯",
            "decision_tree": "决策树 C4.5", "svm": "SVM",
            "kmeans": "K-Means", "mlp": "MLP",
        }

        result = {
            "dataset": ds_key,
            "algorithm": algo_key,
            "algo_name": algo_names.get(algo_key, algo_key),
            "params": params,
        }

        if algo_key == "kmeans":
            # K-Means 聚类评估
            n_clusters = int(params.get("n_clusters", 3))
            t0 = time.perf_counter()
            km = _get_algo_instance("kmeans", params)
            km.fit(X)
            elapsed = time.perf_counter() - t0

            sil = KMeans.silhouette_score(X, km.labels_) if len(X) <= 2000 else None
            result.update({
                "accuracy": round(sil, 4) if sil else None,
                "inertia": round(float(km.inertia_), 4),
                "silhouette": round(sil, 4) if sil else None,
                "train_time": round(elapsed, 4),
                "type": "clustering",
                "n_clusters": n_clusters,
            })
        else:
            # 分类器训练
            X_tr, X_te, y_tr, y_te = train_test_split(
                X, y, test_size=test_size, random_state=42
            )

            t0 = time.perf_counter()
            model.fit(X_tr, y_tr)
            y_pred = model.predict(X_te)
            elapsed = time.perf_counter() - t0

            metrics = compute_all_metrics(y_te, y_pred)

            result.update({
                "accuracy": round(metrics.get("accuracy", 0), 4),
                "precision": round(metrics.get("precision_macro", metrics.get("precision_weighted", 0)), 4),
                "recall": round(metrics.get("recall_macro", metrics.get("recall_weighted", 0)), 4),
                "f1": round(metrics.get("f1_macro", metrics.get("f1_weighted", 0)), 4),
                "train_time": round(elapsed, 4),
                "test_samples": len(y_te),
                "type": "classification",
            })

            # 生成混淆矩阵图
            cm = metrics.get("confusion_matrix")
            if cm is not None:
                chart_b64 = _plot_confusion_matrix(cm, cnames)
                result["confusion_matrix_b64"] = chart_b64

            # ── 新增：多种评估可视化 ──
            # 1. 每类指标柱状图
            try:
                result["per_class_b64"] = _plot_per_class_metrics(y_te, y_pred, cnames)
            except Exception:
                pass

            # 2. PCA 2D 分布（预测标签着色）
            try:
                result["pca_pred_b64"] = _plot_pca_prediction(X_te, y_te, y_pred, cnames)
            except Exception:
                pass

            # 3. 预测类别分布饼图
            try:
                result["pred_distribution_b64"] = _plot_pred_distribution(y_te, y_pred, cnames)
            except Exception:
                pass

            # 交叉验证
            if use_cv:
                cv_result = cross_validate(
                    lambda: _get_algo_instance(algo_key, params),
                    X, y, k=5, n_runs=1
                )
                result.update({
                    "cv_accuracy": round(cv_result.get("accuracy_mean", 0), 4),
                    "cv_accuracy_std": round(cv_result.get("accuracy_std", 0), 4),
                    "cv_f1": round(cv_result.get("f1_macro_mean", 0), 4),
                })

        # 缓存结果
        run_id = str(_run_counter[0])
        _run_counter[0] += 1
        result["run_id"] = run_id
        results_cache[run_id] = result

        return jsonify(result)

    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


def _plot_confusion_matrix(cm, class_names):
    """生成混淆矩阵 base64 图片"""
    fig, ax = plt.subplots(figsize=(6, 5))
    im = ax.imshow(cm, cmap="YlOrRd", aspect="auto")

    n = cm.shape[0]
    for i in range(n):
        for j in range(n):
            ax.text(j, i, int(cm[i, j]), ha="center", va="center",
                    color="white" if cm[i, j] > cm.max() / 2 else "black",
                    fontsize=10, fontweight="bold")

    labels = class_names if class_names else [str(i) for i in range(n)]
    ax.set_xticks(range(n))
    ax.set_yticks(range(n))
    ax.set_xticklabels(labels, rotation=45, ha="right")
    ax.set_yticklabels(labels)
    ax.set_xlabel("预测类别")
    ax.set_ylabel("真实类别")
    ax.set_title("混淆矩阵", fontweight="bold")
    plt.colorbar(im, ax=ax, shrink=0.8)
    fig.tight_layout()

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=100, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode("utf-8")


def _plot_per_class_metrics(y_true, y_pred, class_names):
    """每类 Precision / Recall / F1 分组柱状图 — 直接从混淆矩阵计算每类指标"""
    # 确保是 numpy 数组
    y_true = np.asarray(y_true)
    y_pred = np.asarray(y_pred)

    n_classes = int(max(y_true.max(), y_pred.max())) + 1
    if n_classes > 20:
        return None  # 类别太多不画

    # ── 直接算混淆矩阵 + 每类指标（不依赖 compute_all_metrics） ──
    cm = np.zeros((n_classes, n_classes), dtype=np.int64)
    for t, p in zip(y_true.astype(np.int64), y_pred.astype(np.int64)):
        cm[t, p] += 1

    tp = np.diag(cm).astype(np.float64)
    fp = cm.sum(axis=0).astype(np.float64) - tp
    fn = cm.sum(axis=1).astype(np.float64) - tp

    eps = 1e-12
    prec = np.where(tp + fp > eps, tp / (tp + fp), 0.0)
    rec  = np.where(tp + fn > eps, tp / (tp + fn), 0.0)
    f1   = np.where(prec + rec > eps, 2.0 * prec * rec / (prec + rec), 0.0)
    # 转为 Python list 用于 matplotlib
    prec = prec.tolist()
    rec  = rec.tolist()
    f1   = f1.tolist()

    labels = class_names if class_names and len(class_names) >= n_classes \
             else [f"Class {i}" for i in range(n_classes)]

    fig, ax = plt.subplots(figsize=(max(6, n_classes*1.2), 5))
    x = np.arange(n_classes)
    width = 0.25

    bars1 = ax.bar(x - width, prec, width, label="Precision", color="#0d9488", alpha=0.85)
    bars2 = ax.bar(x, rec, width, label="Recall", color="#d97706", alpha=0.85)
    bars3 = ax.bar(x + width, f1, width, label="F1", color="#7c3aed", alpha=0.85)

    # 标注数值
    for bars in [bars1, bars2, bars3]:
        for bar in bars:
            h = bar.get_height()
            if h > 0.05:
                ax.text(bar.get_x() + bar.get_width()/2., h + 0.02,
                        f'{h:.2f}', ha='center', va='bottom', fontsize=7)

    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=30, ha='right', fontsize=9)
    ax.set_ylabel("Score", fontsize=11)
    ax.set_title("每类评估指标对比", fontweight="bold", fontsize=13)
    ax.legend(fontsize=9)
    ax.set_ylim(0, 1.15)
    ax.grid(axis='y', alpha=0.3)
    fig.tight_layout()

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=120, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode("utf-8")


def _plot_pca_prediction(X, y_true, y_pred, class_names):
    """PCA 2D 降维，展示预测 vs 真实标签分布"""
    from sklearn.decomposition import PCA

    if X.shape[1] < 2:
        return None

    pca = PCA(n_components=2)
    X_2d = pca.fit_transform(X)

    n_classes = len(np.unique(y_true))
    labels = class_names if class_names and len(class_names) >= n_classes \
             else [f"Class {i}" for i in range(n_classes)]
    colors = plt.cm.tab10(np.linspace(0, 1, max(n_classes, 3)))

    fig, axes = plt.subplots(1, 2, figsize=(12, 5))

    # 左：真实标签
    ax = axes[0]
    for c in range(n_classes):
        mask = y_true == c
        ax.scatter(X_2d[mask, 0], X_2d[mask, 1], c=[colors[c]], label=labels[c],
                   alpha=0.6, s=20, edgecolors='none')
    ax.set_title("真实标签分布", fontweight="bold")
    ax.set_xlabel(f"PC1 ({pca.explained_variance_ratio_[0]*100:.1f}%)")
    ax.set_ylabel(f"PC2 ({pca.explained_variance_ratio_[1]*100:.1f}%)")
    if n_classes <= 10:
        ax.legend(fontsize=7, markerscale=0.8)

    # 右：预测标签，标注错误点
    ax = axes[1]
    correct = y_true == y_pred
    wrong = ~correct
    # 先画正确的
    for c in range(n_classes):
        mask = (y_true == c) & correct
        if mask.any():
            ax.scatter(X_2d[mask, 0], X_2d[mask, 1], c=[colors[c]],
                       alpha=0.5, s=18, edgecolors='none')
    # 再画错误的（用 × 标记）
    if wrong.any():
        ax.scatter(X_2d[wrong, 0], X_2d[wrong, 1], c='#dc2626',
                   marker='x', s=40, alpha=0.8, linewidths=1.5, label='预测错误')
    ax.set_title("预测结果分布 (✕ = 错误)", fontweight="bold")
    ax.set_xlabel(f"PC1 ({pca.explained_variance_ratio_[0]*100:.1f}%)")
    ax.set_ylabel(f"PC2 ({pca.explained_variance_ratio_[1]*100:.1f}%)")
    if wrong.any():
        ax.legend(fontsize=7)

    fig.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=120, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode("utf-8")


def _plot_pred_distribution(y_true, y_pred, class_names):
    """预测类别分布对比饼图"""
    n_classes = len(np.unique(np.concatenate([y_true, y_pred])))
    if n_classes > 15:
        return None

    labels = class_names if class_names and len(class_names) >= n_classes \
             else [f"Class {i}" for i in range(n_classes)]
    colors = plt.cm.Set3(np.linspace(0, 1, n_classes))

    fig, axes = plt.subplots(1, 2, figsize=(10, 4.5))

    # 真实分布
    true_counts = [int(np.sum(y_true == c)) for c in range(n_classes)]
    axes[0].pie(true_counts, labels=labels, autopct='%1.1f%%',
                colors=colors, startangle=90, textprops={'fontsize': 8})
    axes[0].set_title("真实类别分布", fontweight="bold")

    # 预测分布
    pred_counts = [int(np.sum(y_pred == c)) for c in range(n_classes)]
    axes[1].pie(pred_counts, labels=labels, autopct='%1.1f%%',
                colors=colors, startangle=90, textprops={'fontsize': 8})
    axes[1].set_title("预测类别分布", fontweight="bold")

    fig.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=120, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode("utf-8")


# ============================================================
# 结果 API
# ============================================================

@app.route("/api/results/<run_id>")
def api_result(run_id):
    """获取缓存的运行结果"""
    if run_id in results_cache:
        return jsonify(results_cache[run_id])
    return jsonify({"error": "结果不存在"}), 404


@app.route("/api/comparison", methods=["POST"])
def api_comparison():
    """多算法对比"""
    data = request.get_json() or {}
    ds_key = data.get("dataset", "iris")
    algorithms = data.get("algorithms", ["knn", "naive_bayes"])
    params_list = data.get("params_list", {})

    results = []
    for algo_key in algorithms:
        try:
            params = params_list.get(algo_key, {})
            resp = api_train_internal(ds_key, algo_key, params)
            results.append(resp)
        except Exception as e:
            results.append({"algorithm": algo_key, "error": str(e)})

    return jsonify({"dataset": ds_key, "results": results})


def api_train_internal(ds_key, algo_key, params):
    """内部训练函数"""
    X, y, fnames, cnames = load_dataset(ds_key)
    scaler = StandardScaler()
    X = scaler.fit_transform(X)

    model = _get_algo_instance(algo_key, params)
    if model is None:
        return {"algorithm": algo_key, "error": "未知算法"}

    X_tr, X_te, y_tr, y_te = train_test_split(X, y, test_size=0.2, random_state=42)
    t0 = time.perf_counter()
    model.fit(X_tr, y_tr)
    y_pred = model.predict(X_te)
    elapsed = time.perf_counter() - t0

    metrics = compute_all_metrics(y_te, y_pred)
    return {
        "algorithm": algo_key,
        "accuracy": round(metrics.get("accuracy", 0), 4),
        "precision": round(metrics.get("precision_macro", 0), 4),
        "recall": round(metrics.get("recall_macro", 0), 4),
        "f1": round(metrics.get("f1_macro", 0), 4),
        "train_time": round(elapsed, 4),
    }


# ============================================================
# 数据上传 API
# ============================================================

def _detect_delimiter(first_line):
    """自动检测 CSV 分隔符"""
    candidates = [',', '\t', ';', '|']
    best, best_count = ',', 0
    for delim in candidates:
        cnt = first_line.count(delim)
        if cnt > best_count:
            best, best_count = delim, cnt
    return best


def _parse_csv(filepath):
    """解析 CSV 文件，自动检测分隔符和表头"""
    with open(filepath, 'r', encoding='utf-8-sig') as f:
        first_line = f.readline().strip()
        f.seek(0)
        delimiter = _detect_delimiter(first_line)
        reader = csv.reader(f, delimiter=delimiter)

        # 尝试读取第一行作为表头
        rows = list(reader)
        if not rows:
            raise ValueError("文件为空")

    # 判断第一行是否为表头（包含非数字的字符串 → 可能是表头）
    first_row = rows[0]
    has_header = any(not (v.strip().replace('.', '').replace('-', '').replace('e', '').replace('E', '')
                          .isdigit() or v.strip() == '')
                     for v in first_row if v.strip())

    if has_header:
        header = [h.strip() for h in first_row]
        data_rows = rows[1:]
    else:
        header = [f"feature_{i}" for i in range(len(first_row))]
        data_rows = rows

    # 转换为 numpy 数组
    data = []
    for row in data_rows:
        try:
            data.append([float(v.strip()) if v.strip() else np.nan for v in row])
        except ValueError:
            # 包含非数字 → 可能是标签/类别列
            data.append(row)

    # 尝试转数值
    try:
        arr = np.array(data, dtype=np.float64)
    except ValueError:
        # 混合类型 → 区分数值列和标签列
        n_cols = len(data[0])
        numeric_cols = []
        label_cols = []
        for j in range(n_cols):
            all_num = True
            for row in data:
                try:
                    float(str(row[j]).strip()) if str(row[j]).strip() else None
                except ValueError:
                    all_num = False
                    break
            if all_num:
                numeric_cols.append(j)
            else:
                label_cols.append(j)

        # 提取数值特征
        numeric_data = [[float(str(row[j]).strip()) if str(row[j]).strip() and j in numeric_cols else np.nan
                        for j in numeric_cols] for row in data]
        arr = np.array(numeric_data, dtype=np.float64)

        # 取最后一列非数值的作为标签
        y = None
        if label_cols:
            label_idx = label_cols[-1]
            y_raw = [str(row[label_idx]).strip() for row in data]
            unique_labels = list(dict.fromkeys(y_raw))  # 保持顺序去重
            label_map = {lbl: i for i, lbl in enumerate(unique_labels)}
            y = np.array([label_map[lbl] for lbl in y_raw])
        header = [header[j] for j in numeric_cols]

        return arr, y, header

    # 纯数值 → 默认最后一列为目标
    if arr.shape[1] > 1:
        X = arr[:, :-1]
        y = arr[:, -1].astype(int)
        feature_names = header[:-1]
    else:
        X = arr
        y = np.zeros(arr.shape[0], dtype=int)
        feature_names = header

    return X, y, feature_names


@app.route("/api/upload", methods=["POST"])
def api_upload():
    """上传 CSV/Excel 文件"""
    if "file" not in request.files:
        return jsonify({"error": "未选择文件"}), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "文件名为空"}), 400

    # 保存文件
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in [".csv", ".tsv", ".txt", ".xlsx", ".xls"]:
        return jsonify({"error": f"不支持的文件格式: {ext}，请上传 CSV/Excel 文件"}), 400

    upload_id = str(uuid.uuid4())[:8]
    safe_name = hashlib.md5(file.filename.encode()).hexdigest()[:8]
    save_path = os.path.join(UPLOAD_FOLDER, f"{upload_id}_{safe_name}{ext}")
    file.save(save_path)

    try:
        if ext in [".xlsx", ".xls"]:
            # Excel 文件
            try:
                import openpyxl
                wb = openpyxl.load_workbook(save_path, data_only=True)
                ws = wb.active
                rows = [[cell.value if cell.value is not None else "" for cell in row]
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                                 max_col=ws.max_column)]
                # 保存为临时 CSV 再解析
                tmp_csv = save_path + ".tmp.csv"
                with open(tmp_csv, 'w', encoding='utf-8-sig', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerows(rows)
                X, y, fnames = _parse_csv(tmp_csv)
                os.remove(tmp_csv)
            except ImportError:
                return jsonify({"error": "需安装 openpyxl 库以支持 Excel 文件: pip install openpyxl"}), 500
        else:
            X, y, fnames = _parse_csv(save_path)

        n_samples, n_features = X.shape
        n_missing = int(np.sum(np.isnan(X)))

        # 若无标签列，生成虚拟标签
        if y is None:
            y = np.zeros(n_samples, dtype=int)

        n_classes = len(np.unique(y))

        # 缓存数据
        upload_cache[upload_id] = {
            "id": upload_id,
            "filename": file.filename,
            "X": X,
            "y": y,
            "feature_names": fnames,
            "class_names": [f"Class_{c}" for c in np.unique(y)],
            "n_samples": n_samples,
            "n_features": n_features,
            "n_classes": n_classes,
            "n_missing": n_missing,
            "save_path": save_path,
            "timestamp": time.time(),
        }

        # 基本统计
        stats = []
        for i in range(min(n_features, 15)):
            col = X[:, i]
            fname = fnames[i] if i < len(fnames) else f"feature_{i}"
            stats.append({
                "index": i, "name": fname,
                "mean": round(float(np.nanmean(col)), 4),
                "std": round(float(np.nanstd(col)), 4),
                "min": round(float(np.nanmin(col)), 4),
                "q25": round(float(np.nanpercentile(col, 25)), 4),
                "median": round(float(np.nanmedian(col)), 4),
                "q75": round(float(np.nanpercentile(col, 75)), 4),
                "max": round(float(np.nanmax(col)), 4),
                "missing": int(np.sum(np.isnan(col))),
            })

        # 类别分布
        unique, counts = np.unique(y, return_counts=True)
        class_dist = {str(c): int(cnt) for c, cnt in zip(unique, counts)}

        return jsonify({
            "upload_id": upload_id,
            "filename": file.filename,
            "n_samples": n_samples,
            "n_features": n_features,
            "n_classes": n_classes,
            "n_missing": n_missing,
            "feature_stats": stats,
            "class_distribution": class_dist,
            "status": "ok",
        })

    except Exception as e:
        traceback.print_exc()
        # 清理临时文件
        if os.path.exists(save_path):
            os.remove(save_path)
        return jsonify({"error": f"文件解析失败: {str(e)}"}), 400


@app.route("/api/upload/<upload_id>/analyze")
def api_upload_analyze(upload_id):
    """对上传数据进行统计分析"""
    if upload_id not in upload_cache:
        return jsonify({"error": "上传数据不存在或已过期"}), 404

    data = upload_cache[upload_id]
    X, y = data["X"], data["y"]
    fnames = data["feature_names"]
    n_features = X.shape[1]

    # 完整统计
    full_stats = []
    for i in range(n_features):
        col = X[:, i]
        fname = fnames[i] if i < len(fnames) else f"feature_{i}"
        full_stats.append({
            "index": i, "name": fname,
            "mean": round(float(np.nanmean(col)), 4),
            "std": round(float(np.nanstd(col)), 4),
            "min": round(float(np.nanmin(col)), 4),
            "q25": round(float(np.nanpercentile(col, 25)), 4),
            "median": round(float(np.nanmedian(col)), 4),
            "q75": round(float(np.nanpercentile(col, 75)), 4),
            "max": round(float(np.nanmax(col)), 4),
            "missing": int(np.sum(np.isnan(col))),
            "skewness": round(float(_safe_skew(col)), 4),
            "kurtosis": round(float(_safe_kurtosis(col)), 4),
        })

    # 相关性矩阵（Top 热力数据）
    corr = np.corrcoef(X[:, :min(n_features, 20)], rowvar=False)
    corr_flat = []
    for i in range(corr.shape[0]):
        for j in range(i + 1, corr.shape[1]):
            if abs(corr[i, j]) > 0.3:  # 只保留显著相关
                corr_flat.append({
                    "feat1": fnames[i] if i < len(fnames) else f"f{i}",
                    "feat2": fnames[j] if j < len(fnames) else f"f{j}",
                    "correlation": round(float(corr[i, j]), 4),
                })

    return jsonify({
        "upload_id": upload_id,
        "n_samples": data["n_samples"],
        "n_features": n_features,
        "n_missing": data["n_missing"],
        "feature_stats": full_stats,
        "top_correlations": sorted(corr_flat, key=lambda x: -abs(x["correlation"]))[:15],
    })


def _safe_skew(col):
    """安全计算偏度"""
    col = col[~np.isnan(col)]
    if len(col) < 3:
        return 0.0
    n = len(col)
    mean = np.mean(col)
    std = np.std(col)
    if std < 1e-10:
        return 0.0
    return (n / ((n - 1) * (n - 2))) * np.sum(((col - mean) / std) ** 3)


def _safe_kurtosis(col):
    """安全计算峰度"""
    col = col[~np.isnan(col)]
    if len(col) < 4:
        return 0.0
    n = len(col)
    mean = np.mean(col)
    std = np.std(col)
    if std < 1e-10:
        return 0.0
    return (n * (n + 1) / ((n - 1) * (n - 2) * (n - 3))) * \
           np.sum(((col - mean) / std) ** 4) - 3 * (n - 1) ** 2 / ((n - 2) * (n - 3))


@app.route("/api/upload/<upload_id>/preprocess", methods=["POST"])
def api_upload_preprocess(upload_id):
    """对上传数据执行预处理"""
    if upload_id not in upload_cache:
        return jsonify({"error": "上传数据不存在"}), 404

    data = upload_cache[upload_id]
    X, y = data["X"], data["y"]

    body = request.get_json() or {}
    handle_missing = body.get("handle_missing", "none")  # none / drop / mean / median
    normalize = body.get("normalize", "none")             # none / standard / minmax
    test_size = float(body.get("test_size", 0.0))         # >0 时执行划分

    # 缺失值处理
    if handle_missing == "drop":
        mask = ~np.any(np.isnan(X), axis=1)
        X = X[mask]
        y = y[mask]
    elif handle_missing == "mean":
        col_means = np.nanmean(X, axis=0)
        X = np.where(np.isnan(X), col_means, X)
    elif handle_missing == "median":
        col_medians = np.nanmedian(X, axis=0)
        X = np.where(np.isnan(X), col_medians, X)

    # 标准化
    if normalize == "standard":
        scaler = StandardScaler()
        X = scaler.fit_transform(X)
    elif normalize == "minmax":
        scaler = MinMaxScaler()
        X = scaler.fit_transform(X)

    # 更新缓存
    upload_cache[upload_id]["X"] = X
    upload_cache[upload_id]["y"] = y
    upload_cache[upload_id]["n_samples"] = X.shape[0]
    upload_cache[upload_id]["n_missing"] = int(np.sum(np.isnan(X)))

    # 划分训练/测试集
    split_info = None
    if test_size > 0:
        X_tr, X_te, y_tr, y_te = train_test_split(X, y, test_size=test_size, random_state=42)
        upload_cache[upload_id]["X_train"] = X_tr
        upload_cache[upload_id]["X_test"] = X_te
        upload_cache[upload_id]["y_train"] = y_tr
        upload_cache[upload_id]["y_test"] = y_te
        split_info = {"train_samples": len(y_tr), "test_samples": len(y_te)}

    return jsonify({
        "upload_id": upload_id,
        "n_samples_after": X.shape[0],
        "n_missing_after": upload_cache[upload_id]["n_missing"],
        "handle_missing": handle_missing,
        "normalize": normalize,
        "split": split_info,
        "status": "ok",
    })


@app.route("/api/upload/<upload_id>/chart/<chart_type>")
def api_upload_chart(upload_id, chart_type):
    """按需生成单个可视化图表，返回 base64"""
    if upload_id not in upload_cache:
        return jsonify({"error": "上传数据不存在"}), 404

    data = upload_cache[upload_id]
    X, y = data["X"], data["y"]
    fnames = data["feature_names"]
    cnames = data["class_names"]
    ds_name = data.get("filename", "upload").rsplit(".", 1)[0]

    charts = {
        "histogram": lambda: plot_histograms(X, fnames, y, cnames, ds_name),
        "boxplot": lambda: plot_boxplots(X, fnames, ds_name),
        "violin": lambda: plot_violin(X, fnames, y, cnames, ds_name),
        "swarm": lambda: plot_swarm(X, fnames, y, cnames, ds_name),
        "heatmap": lambda: plot_correlation_heatmap(X, fnames, ds_name),
        "scatter": lambda: plot_scatter_matrix(X, fnames, y, cnames, ds_name),
        "pca2d": lambda: plot_pca_2d(X, y, cnames, ds_name)[0],
        "pca3d": lambda: plot_pca_3d(X, y, cnames, ds_name)[0],
        "pca_var": lambda: plot_pca_explained_variance(X, ds_name),
        "parallel": lambda: plot_parallel_coordinates(X, fnames, y, cnames, ds_name),
        "radar": lambda: plot_radar(X, y, fnames, cnames, ds_name),
        "dendrogram": lambda: plot_dendrogram(X, fnames, ds_name),
        "ecdf": lambda: plot_ecdf(X, fnames, ds_name),
        "pie": lambda: plot_class_pie(y, cnames, ds_name),
        "pair_density": lambda: plot_pair_density(X, fnames, y, cnames, ds_name),
    }

    if chart_type not in charts:
        return jsonify({"error": f"未知图表类型: {chart_type}，可选: {list(charts.keys())}"}), 400

    try:
        path = charts[chart_type]()
        # 读取生成的图片并转 base64
        with open(path, 'rb') as f:
            b64 = base64.b64encode(f.read()).decode('utf-8')
        return jsonify({
            "chart_type": chart_type,
            "image_b64": b64,
            "format": "png",
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/upload/<upload_id>/chart/all")
def api_upload_all_charts(upload_id):
    """批量生成所有图表"""
    chart_types = [
        "histogram", "boxplot", "violin", "heatmap", "pie",
        "scatter", "pca2d", "pca_var", "ecdf", "parallel",
    ]
    results = {}
    for ct in chart_types:
        try:
            # 调用内部函数而非 HTTP 路由
            resp = _generate_chart_internal(upload_id, ct)
            results[ct] = resp
        except Exception as e:
            results[ct] = {"error": str(e)}

    return jsonify({"upload_id": upload_id, "charts": results})


def _generate_chart_internal(upload_id, chart_type):
    """内部图表生成（不经过 HTTP）"""
    if upload_id not in upload_cache:
        raise ValueError("上传数据不存在")

    data = upload_cache[upload_id]
    X, y = data["X"], data["y"]
    fnames = data["feature_names"]
    cnames = data["class_names"]
    ds_name = data.get("filename", "upload").rsplit(".", 1)[0]

    charts = {
        "histogram": lambda: plot_histograms(X, fnames, y, cnames, ds_name),
        "boxplot": lambda: plot_boxplots(X, fnames, ds_name),
        "violin": lambda: plot_violin(X, fnames, y, cnames, ds_name),
        "swarm": lambda: plot_swarm(X, fnames, y, cnames, ds_name),
        "heatmap": lambda: plot_correlation_heatmap(X, fnames, ds_name),
        "scatter": lambda: plot_scatter_matrix(X, fnames, y, cnames, ds_name),
        "pca2d": lambda: plot_pca_2d(X, y, cnames, ds_name)[0],
        "pca3d": lambda: plot_pca_3d(X, y, cnames, ds_name)[0],
        "pca_var": lambda: plot_pca_explained_variance(X, ds_name),
        "parallel": lambda: plot_parallel_coordinates(X, fnames, y, cnames, ds_name),
        "radar": lambda: plot_radar(X, y, fnames, cnames, ds_name),
        "dendrogram": lambda: plot_dendrogram(X, fnames, ds_name),
        "ecdf": lambda: plot_ecdf(X, fnames, ds_name),
        "pie": lambda: plot_class_pie(y, cnames, ds_name),
        "pair_density": lambda: plot_pair_density(X, fnames, y, cnames, ds_name),
    }

    if chart_type not in charts:
        raise ValueError(f"未知图表类型: {chart_type}")

    path = charts[chart_type]()
    with open(path, 'rb') as f:
        b64 = base64.b64encode(f.read()).decode('utf-8')
    return {"chart_type": chart_type, "image_b64": b64, "format": "png"}


@app.route("/api/upload/<upload_id>/report")
def api_upload_report(upload_id):
    """生成并下载 HTML 数据分析报告"""
    if upload_id not in upload_cache:
        return jsonify({"error": "上传数据不存在"}), 404

    data = upload_cache[upload_id]
    X, y = data["X"], data["y"]
    fnames = data["feature_names"]
    cnames = data["class_names"]
    ds_name = data.get("filename", "upload").rsplit(".", 1)[0]

    try:
        html, paths, html_path = generate_html_report(
            ds_name, X, y, fnames, cnames,
            upload_name=data["filename"]
        )

        return jsonify({
            "upload_id": upload_id,
            "report_html": html,          # 完整 HTML 内容
            "report_path": html_path,     # 文件路径
            "chart_count": len(paths),
            "status": "ok",
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/upload/<upload_id>/report/download")
def api_upload_report_download(upload_id):
    """直接下载 HTML 报告文件"""
    if upload_id not in upload_cache:
        return jsonify({"error": "上传数据不存在"}), 404

    data = upload_cache[upload_id]
    X, y = data["X"], data["y"]
    fnames = data["feature_names"]
    cnames = data["class_names"]
    ds_name = data.get("filename", "upload").rsplit(".", 1)[0]

    try:
        html, paths, html_path = generate_html_report(
            ds_name, X, y, fnames, cnames,
            upload_name=data["filename"]
        )
        return send_file(
            html_path,
            mimetype='text/html',
            as_attachment=True,
            download_name=f"{ds_name}_analysis_report.html"
        )
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/upload/list")
def api_upload_list():
    """列出已上传的数据集"""
    result = []
    for uid, data in upload_cache.items():
        result.append({
            "upload_id": uid,
            "filename": data["filename"],
            "n_samples": data["n_samples"],
            "n_features": data["n_features"],
            "n_classes": data["n_classes"],
            "timestamp": data.get("timestamp", 0),
        })
    return jsonify(sorted(result, key=lambda x: -x["timestamp"]))


@app.route("/api/upload/<upload_id>/predict", methods=["POST"])
def api_upload_predict(upload_id):
    """对上传数据使用算法进行预测"""
    if upload_id not in upload_cache:
        return jsonify({"error": "上传数据不存在"}), 404

    body = request.get_json() or {}
    algo_key = body.get("algorithm", "knn")
    params = body.get("params", {})

    data = upload_cache[upload_id]
    X, y = data["X"], data["y"]

    # 预处理（缺失值填充 + 标准化）
    col_means = np.nanmean(X, axis=0)
    X_clean = np.where(np.isnan(X), col_means, X)

    scaler = StandardScaler()
    X_clean = scaler.fit_transform(X_clean)

    model = _get_algo_instance(algo_key, params)
    if model is None:
        return jsonify({"error": f"未知算法: {algo_key}"}), 400

    X_tr, X_te, y_tr, y_te = train_test_split(
        X_clean, y, test_size=0.2, random_state=42
    )

    t0 = time.perf_counter()
    model.fit(X_tr, y_tr)
    y_pred = model.predict(X_te)
    elapsed = time.perf_counter() - t0

    metrics = compute_all_metrics(y_te, y_pred)

    # 混淆矩阵图
    cm = metrics.get("confusion_matrix")
    cm_b64 = None
    if cm is not None:
        fig, ax = plt.subplots(figsize=(5, 4))
        im = ax.imshow(cm, cmap="YlOrRd", aspect="auto")
        n = cm.shape[0]
        for i in range(n):
            for j in range(n):
                ax.text(j, i, int(cm[i, j]), ha="center", va="center",
                        color="white" if cm[i, j] > cm.max() / 2 else "black",
                        fontsize=9, fontweight="bold")
        ax.set_xlabel("预测类别"); ax.set_ylabel("真实类别")
        ax.set_title("混淆矩阵", fontweight="bold")
        plt.colorbar(im, ax=ax, shrink=0.8)
        fig.tight_layout()
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=100, bbox_inches="tight")
        plt.close(fig)
        buf.seek(0)
        cm_b64 = base64.b64encode(buf.read()).decode("utf-8")

    return jsonify({
        "upload_id": upload_id,
        "algorithm": algo_key,
        "accuracy": round(metrics.get("accuracy", 0), 4),
        "precision": round(metrics.get("precision_macro", metrics.get("precision_weighted", 0)), 4),
        "recall": round(metrics.get("recall_macro", metrics.get("recall_weighted", 0)), 4),
        "f1": round(metrics.get("f1_macro", metrics.get("f1_weighted", 0)), 4),
        "train_time": round(elapsed, 4),
        "test_samples": len(y_te),
        "confusion_matrix_b64": cm_b64,
        # ── 新增图表 ──
        "per_class_b64": _try_chart(lambda: _plot_per_class_metrics(y_te, y_pred,
            data.get("class_names", [f"C{i}" for i in range(len(np.unique(y_te)))]))),
        "pca_pred_b64": _try_chart(lambda: _plot_pca_prediction(X_te, y_te, y_pred,
            data.get("class_names", [f"C{i}" for i in range(len(np.unique(y_te)))]))),
        "pred_distribution_b64": _try_chart(lambda: _plot_pred_distribution(y_te, y_pred,
            data.get("class_names", [f"C{i}" for i in range(len(np.unique(y_te)))]))),
    })

def _try_chart(fn):
    """安全调用图表生成，失败返回 None"""
    try:
        return fn()
    except Exception:
        return None


def create_app():
    os.makedirs(WEB_CONFIG["upload_folder"], exist_ok=True)
    return app


if __name__ == "__main__":
    create_app()
    app.run(
        host=WEB_CONFIG["host"],
        port=WEB_CONFIG["port"],
        debug=WEB_CONFIG["debug"],
    )
